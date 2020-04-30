from flask import Flask, render_template
from flask_ask import Ask, question, statement
from random import randint
from openpyxl import load_workbook
wb = load_workbook('Savvy.xlsx', data_only=True) 
ws = wb.active 
app = Flask(__name__)

ask = Ask(app, "/alexa")

@ask.launch
def launched():
    return question('Welcome aboard the Black pearl! What be your favourite character, sire?')

@ask.intent('HectorBarbossa')
def getHector():
    return statement(ws.cell(randint(3,ws.cell(1,1).value+2),1).value)

@ask.intent('CaptainJackSparrow')
def getJack():
    return statement(ws.cell(randint(3,ws.cell(1,2).value+2),2).value)

@ask.intent('ElizabethSwann')
def getEli():
    return statement(ws.cell(randint(3,ws.cell(1,3).value+2),3).value)

@ask.intent('WillTurner')
def getWill():
    return statement(ws.cell(randint(3,ws.cell(1,4).value+2),4).value)

@ask.intent('DavyJones')
def getJones():
    return statement(ws.cell(randint(3,ws.cell(1,5).value+2),5).value)

@ask.intent('JoshameeGibbs')
def getGibbs():
    return statement(ws.cell(randint(3,ws.cell(1,6).value+2),6).value)

# Default intent Trigger for Help
@ask.intent('AMAZON.HelpIntent')
def help():
    help_text = render_template('help')
    return question(help_text).reprompt(help_text)

# Default intent for Stop
@ask.intent('AMAZON.StopIntent')
def stop():
    bye_text = render_template('bye')
    return statement(bye_text)

# Default intent for Cancel
@ask.intent('AMAZON.CancelIntent')
def cancel():
    bye_text = render_template('bye')
    return statement(bye_text)


@app.route('/', methods=["GET", "POST"])
def hello():
    return "Aye, Aye Cap'n!"


if __name__ == '__main__':
    app.run(debug=True)

